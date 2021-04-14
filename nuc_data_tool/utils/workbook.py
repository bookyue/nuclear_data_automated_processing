from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def get_column_index(filename):
    """
    :param filename: input file
    :return: data rang of column ex: (2, 32)
    """
    start_column_index = 2
    wb = load_workbook(filename)
    sheet = wb.worksheets[0]
    # print(sheet)
    max_column = sheet.max_column
    # print(max_column)
    return start_column_index, max_column


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters
    ----------
    filename : Path or str or pd.ExcelWriter
        File path or existing ExcelWriter
        (Example: '/path/to/file.xlsx')
    df : pd.DataFrame
        DataFrame to save to workbook
    sheet_name : str
        Name of sheet which will contain DataFrame.
        (default: 'Sheet1')
    startrow : int
        upper left cell row to dump data frame.
        Per default (startrow=None) calculate the last row
        in the existing DF and write to the next row...
    truncate_sheet : bool
        truncate (remove and recreate) [sheet_name]
        before writing DataFrame to Excel file
    to_excel_kwargs : kwarg
        arguments which will be passed to `DataFrame.to_excel()`
        [can be a dictionary]
    Returns
    -------

    """
    # Excel file doesn't exist - saving and exiting
    if not Path(filename).is_file():
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def save_to_excel(dict_df, file_name, dir_path):
    """
    保存结果至xlsx文件
    keys of dict 为 sheet name
    values of dict 为 worksheet table

    Parameters
    ----------
    dict_df : dict[str, pd.DataFrame]
    file_name : str
    dir_path : Path or str

    Returns
    -------
    """

    dir_path = Path(dir_path)
    dir_path.mkdir(parents=True, exist_ok=True)
    file_path = dir_path.joinpath(file_name)

    for key in dict_df:
        append_df_to_excel(file_path,
                           dict_df[key],
                           sheet_name=key,
                           index=False,
                           encoding='utf-8')
